# -*- coding: utf-8 -*-

# Importações
from tkinter import *
import sqlite3
import openpyxl as opxl
import os

# Caminho do arquivo
script_directory = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(script_directory, 'gerenciamento.xlsx')

# Caminho do banco de dados
diretorio_atual = os.path.dirname(__file__)
database_file_path = os.path.join(diretorio_atual, 'gerenciamentoitens.sqlite')

# Openpyxl e Excel
arquivo = opxl.load_workbook(excel_file_path)
sheet = arquivo.active

# Funções dos Botões
def cadastrar():
    # Conexão com banco de dados
    conexao = sqlite3.connect(database_file_path)
    cursor = conexao.cursor()

    try:
        nome_item = entryNome.get().title()
        quantidade = int(entryQuantidade.get())
        preco = float(entryPreco.get())

        # Banco de Dados
        comando_verificacao = """SELECT COUNT(*) FROM itens WHERE nome_item = ?;"""
        cursor.execute(comando_verificacao, (nome_item,))
        quantidade_registros = cursor.fetchone()[0]

        if quantidade_registros > 0:
            comando_somar = """UPDATE itens
                            SET quantidade = quantidade + ?,
                                preco = ?
                            WHERE "nome_item" = ?;
                            """
            cursor.execute(comando_somar, (quantidade, preco, nome_item))
            conexao.commit()

            # Excel
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] == nome_item:
                    quantidade_nova = quantidade
                    preco_novo = preco
                    sheet.cell(row=index, column=2, value=quantidade_nova)
                    sheet.cell(row=index, column=3, value=preco_novo)
                    arquivo.save(excel_file_path)
                    break
        else:
            comando = f"""INSERT INTO itens ("nome_item", "quantidade", "preco")
                        VALUES (?, ?, ?);"""
            cursor.execute(comando, (nome_item, quantidade, preco))
            conexao.commit()

        # Excel
            row = [nome_item, quantidade, preco]
            sheet.append(row)
            arquivo.save(excel_file_path)
    except ValueError:
        textDetalhes.delete("1.0", END)
        texto = "ERRO: digite um\nvalor compatível."
        textDetalhes.insert("1.0", texto)
    finally:
        cursor.close()
        conexao.close()

def registrarUso():
    # Conexão com banco de dados
    conexao = sqlite3.connect(database_file_path)
    cursor = conexao.cursor()

    # Programa
    try:
        nome_item = entryNome.get().title()
        quantidade = int(entryQuantidade.get())

        # Banco de Dados
        comando_verificacao = """SELECT quantidade FROM itens WHERE nome_item = ?;"""
        cursor.execute(comando_verificacao, (nome_item,))
        quantidade_db = cursor.fetchone()

        if int(quantidade_db[0]) < quantidade:
            textDetalhes.delete("1.0", END)
            texto = "ERRO: digite uma\nquantidade menor do\nque a atual."
            textDetalhes.insert("1.0", texto)
        else:
            comando = f"""UPDATE itens
                        SET quantidade = quantidade - ?
                        WHERE "nome_item" = ?;
                        """
            cursor.execute(comando, (quantidade, nome_item))
            conexao.commit()

            # Excel
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] == nome_item:
                    if row[1] >= quantidade:
                        quantidade_nova = row[1] - quantidade
                        sheet.cell(row=index, column=2, value=quantidade_nova)
                        arquivo.save(excel_file_path)
                        break
    except ValueError:
        textDetalhes.delete("1.0", END)
        texto = "ERRO: digite uma\nquantidade menor do\nque a atual."
        textDetalhes.insert("1.0", texto)
    finally:
        cursor.close()
        conexao.close()

def deletar():
    # Conexão com banco de dados
    conexao = sqlite3.connect(database_file_path)
    cursor = conexao.cursor()

    # Programa
    try:
        nome_item = entryNome.get().title()

        # Banco de Dados
        comando = f"""DELETE FROM itens WHERE "nome_item" = ?;"""
        cursor.execute(comando, (nome_item,))
        conexao.commit()

        # Excel
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == nome_item:
                sheet.delete_rows(index)
                arquivo.save(excel_file_path)
                break
    except ValueError:
        textDetalhes.delete("1.0", END)
        texto = "ERRO: digite um item\nexistente."
        textDetalhes.insert("1.0", texto)
    finally:
        cursor.close()
        conexao.close()

def procurar():
    # Conexão com banco de dados
    conexao = sqlite3.connect(database_file_path)
    cursor = conexao.cursor()

    # Programa
    try:
        nome_item = entryNome.get().title()
        comando = f"""SELECT * FROM Itens WHERE "nome_item" = ?;"""
        cursor.execute(comando, (nome_item,))
        valor = cursor.fetchone()
        textDetalhes.delete("1.0", END)
        texto = f"- Item: {valor[1]}\n- Quantidade: \n{valor[2]} unidades\n- Preço: R$ {valor[3]}"
        textDetalhes.insert("1.0", texto)
    except TypeError:
        textDetalhes.delete("1.0", END)
        texto = "ERRO: digite um\nitem existente."
        textDetalhes.insert("1.0", texto)
    finally:
        cursor.close()
        conexao.close()

# Interface Tkinter
window = Tk()
window.title("Gerenciamento de Itens")

window.geometry("382x484")
window.configure(bg = "#ffffff")
canvas = Canvas(
    window,
    bg = "#ffffff",
    height = 484,
    width = 382,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge")
canvas.place(x = 0, y = 0)

background_img = PhotoImage(file = f"background.png")
background = canvas.create_image(
    191.0, 242.0,
    image=background_img)

img0 = PhotoImage(file = f"img0.png")
botaoCadastrar = Button(
    image = img0,
    borderwidth = 0,
    highlightthickness = 0,
    command =lambda: cadastrar(),
    relief = "flat")

botaoCadastrar.place(
    x = 47, y = 107,
    width = 121,
    height = 31)

img1 = PhotoImage(file = f"img1.png")
botaoDetalhes = Button(
    image = img1,
    borderwidth = 0,
    highlightthickness = 0,
    command = procurar,
    relief = "flat")

botaoDetalhes.place(
    x = 212, y = 150,
    width = 121,
    height = 31)

img2 = PhotoImage(file = f"img2.png")
botaoDeletar = Button(
    image = img2,
    borderwidth = 0,
    highlightthickness = 0,
    command = deletar,
    relief = "flat")

botaoDeletar.place(
    x = 47, y = 150,
    width = 121,
    height = 31)

img3 = PhotoImage(file = f"img3.png")
botaoRegistrarUso = Button(
    image = img3,
    borderwidth = 0,
    highlightthickness = 0,
    command = registrarUso,
    relief = "flat")

botaoRegistrarUso.place(
    x = 212, y = 107,
    width = 121,
    height = 31)

entry0_img = PhotoImage(file = f"img_textBox0.png")
entryPreco_bg = canvas.create_image(
    108.0, 394.5,
    image = entry0_img)

entryPreco = Entry(
    bd = 0,
    bg = "#d9d9d9",
    highlightthickness = 0)

entryPreco.place(
    x = 36, y = 381,
    width = 144,
    height = 25)

canvas.create_text(
    98.5, 365.5,
    text = "Preço (em R$)",
    fill = "#ffffff",
    font = ("Inter-Medium", int(15.0)))

entry1_img = PhotoImage(file = f"img_textBox1.png")
entryQuantidade_bg = canvas.create_image(
    108.0, 329.5,
    image = entry1_img)

entryQuantidade = Entry(
    bd = 0,
    bg = "#d9d9d9",
    highlightthickness = 0)

entryQuantidade.place(
    x = 36, y = 316,
    width = 144,
    height = 25)

entry2_img = PhotoImage(file = f"img_textBox2.png")
entryNome_bg = canvas.create_image(
    108.0, 268.5,
    image = entry2_img)

entryNome = Entry(
    bd = 0,
    bg = "#d9d9d9",
    highlightthickness = 0)

entryNome.place(
    x = 36, y = 255,
    width = 144,
    height = 25)

entry3_img = PhotoImage(file = f"img_textBox3.png")
entry3_bg = canvas.create_image(
    273.0, 330.5,
    image = entry3_img)

textDetalhes = Text(
    bd = 0,
    bg = "#d9d9d9",
    highlightthickness = 0)

textDetalhes.place(
    x = 195, y = 222,
    width = 156,
    height = 215)

canvas.create_text(
    58.5, 242.0,
    text = "Nome",
    fill = "#ffffff",
    font = ("Inter-Medium", int(15.0)))

canvas.create_text(
    82.5, 302.0,
    text = "Quantidade",
    fill = "#ffffff",
    font = ("Inter-Medium", int(15.0)))

canvas.create_text(
    273.0, 208.0,
    text = "Detalhes do Item",
    fill = "#ffffff",
    font = ("Inter-Medium", int(13.0)))

window.resizable(False, False)
window.mainloop()
